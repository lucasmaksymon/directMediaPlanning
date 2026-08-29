"""Parse ADMIN 2026.xlsx into data/admin-2026.json for the ERP import."""
from __future__ import annotations

import json
import re
from datetime import date, datetime
from pathlib import Path

import openpyxl

SRC = Path(r"C:\Users\lucas\Downloads\ADMIN 2026 .xlsx")
OUT = Path(__file__).resolve().parents[1] / "data" / "admin-2026.json"

VENDOR_ALIASES = {
    "OSP": "OSP S.A.",
    "OSP S.A.": "OSP S.A.",
    "PC PULICIDAD": "PC PUBLICIDAD",
    "PC PUBLICIDAD": "PC PUBLICIDAD",
    "WALL STREET SA": "WALL STREET SA",
    "WALL STREET VIA PUBLICA S A": "WALL STREET SA",
    "BYV SRL": "BYV 360 S.R.L.",
    "BYV 360 S.R.L.": "BYV 360 S.R.L.",
    "PENAPAN SA": "PENAPAN",
    "CRISTIAN DE NARDO": "CRISTIAN DE NARDO",
    "DE NARDO CRISTIAN": "CRISTIAN DE NARDO",
    "DE NARDO CRISTIAN ALEJANDRO": "CRISTIAN DE NARDO",
    "GOVP": "GOVP",
    "GOVP S. R. L.": "GOVP",
    "GOVP S.R.L.": "GOVP",
    "VIACART": "VIACART SA",
    "VIACART SA": "VIACART SA",
    "VIAGRAPHIC": "VIAGRAPHIC",
    "VIAGRAPHIC MIDIOS": "VIAGRAPHIC",
    "BROPRINTER": "BROPRINTER",
    "BROPRINTER SRL": "BROPRINTER",
    "MULTIPOSTER": "MULTIPOSTER",
    "MULTIPOSTER S A": "MULTIPOSTER",
    "SENAL VP": "SEÑAL VP",
    "SEÑAL VP": "SEÑAL VP",
    "SEÑAL VP S.A.S.": "SEÑAL VP",
    "SENAL VP S.A.S.": "SEÑAL VP",
}

CLIENT_ALIASES = (
    ("DF SAU", "DF SAU"),
    ("DGE", "DGE"),
    ("DFE", "DF "),
    ("DF ", "DF "),
    ("MIMO", "MIMO"),
    ("MAPSA", "MAPSA"),
    ("SALTAR", "SALTAR"),
    ("APYCE", "APYCE"),
    ("AHYRE", "AHYRE"),
    ("ANUEL", "SERGIO (ANUEL)"),
    ("ALEX", "SERGIO (ALEX"),
    ("GELLY", "SERGIO (GELLY)"),
    ("SERGIO", "SERGIO"),
)


def iso(v):
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, str) and "/" in v:
        try:
            d, m, y = v.split("/")
            year = int(y)
            if year < 100:
                year += 2000
            return f"{year:04d}-{int(m):02d}-{int(d):02d}"
        except Exception:
            return None
    return None


def num(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def txt(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def vendor_name(raw):
    s = txt(raw)
    if not s:
        return None
    key = " ".join(s.upper().replace("Á", "A").replace("É", "E").replace("Í", "I").replace("Ó", "O").replace("Ú", "U").split())
    if key in VENDOR_ALIASES:
        return VENDOR_ALIASES[key]
    if s.upper() in VENDOR_ALIASES:
        return VENDOR_ALIASES[s.upper()]
    return s


def resolve_client(alias, clients, orders, oc_key=None):
    if oc_key and oc_key in orders:
        return orders[oc_key]["client"]
    raw = (alias or "").strip()
    if not raw:
        return None
    if raw in clients:
        return raw
    upper = raw.upper()
    for name in clients:
        if name.upper() == upper or name.upper().startswith(upper + " ") or name.upper().startswith(upper + "("):
            return name
    for prefix, needle in CLIENT_ALIASES:
        if upper == prefix or upper.startswith(prefix):
            for name in clients:
                if needle.upper() in name.upper() or name.upper().startswith(needle.upper()):
                    return name
    return raw


def oc_key(raw):
    if raw is None or raw == "":
        return None
    if isinstance(raw, bool):
        return None
    if isinstance(raw, (int, float)):
        return str(int(raw))
    s = str(raw).strip()
    if not s:
        return None
    if re.fullmatch(r"\d+[.,]\d+", s):
        # 12.38 → orden 12; 07.26 se trata como código de mes
        if re.fullmatch(r"0?\d+[.,]26", s):
            return s.replace(",", ".")
        return str(int(float(s.replace(",", "."))))
    return s


def month_from_key(key, issued):
    if key and re.fullmatch(r"0?\d+\.26", key):
        return int(key.split(".")[0])
    if issued:
        try:
            return int(issued[5:7])
        except Exception:
            pass
    return 1


def parse_receipt_number(raw):
    s = txt(raw)
    if not s:
        return None
    upper = s.upper()
    if "PENDIENTE" in upper or "SALDO A FAVOR" in upper:
        return None
    m = re.search(r"X\s*0*(\d+)", upper)
    if m:
        return int(m.group(1))
    return None


def parse_op_number(raw):
    s = txt(raw)
    if not s:
        return None
    upper = s.upper().strip()
    if upper == "TRA":
        return 165
    m = re.search(r"OP\s*(\d+)", upper)
    if m:
        return int(m.group(1))
    if upper.isdigit():
        return int(upper)
    return None


def parse_comp(raw):
    s = txt(raw)
    if not s:
        return None
    upper = s.upper().replace(".", "")
    credit = upper.startswith("NC")
    m = re.search(r"([A-Z]+)?\s*(\d+)\s*-\s*(\d+)", upper)
    if m:
        kind = m.group(1) or ("NC" if credit else "A")
        return {
            "raw": s,
            "docType": "NC" if credit or kind == "NC" else kind,
            "pos": int(m.group(2)),
            "number": int(m.group(3)),
            "credit": credit or kind == "NC",
        }
    m = re.search(r"(\d+)", upper)
    if not m:
        return None
    return {
        "raw": s,
        "docType": "NC" if credit else "A",
        "pos": 0,
        "number": int(m.group(1)),
        "credit": credit,
    }


def paid(v):
    s = (txt(v) or "").upper()
    if "PAGADO" in s or "COBRADO" in s:
        return 1
    return 0


def fill(prev, val):
    return val if val not in (None, "") else prev


def ars(raw):
    s = str(raw).strip().replace("$", "").replace(" ", "")
    if not s:
        return None
    if "," in s:
        try:
            return float(s.replace(".", "").replace(",", "."))
        except ValueError:
            return None
    try:
        return float(s)
    except ValueError:
        return None


def parse_treasury_notes(raw):
    """Extrae e-cheq y transferencias del texto de PAGO A PROV."""
    s = txt(raw)
    if not s:
        return []
    text = s.replace("\r", "\n")
    found = []
    for m in re.finditer(
        r"(?:ENDOSO?\s+)?ECHEQ\s+(\d+)\s+(?:(\d{1,2}/\d{1,2}/\d{2,4})\s+)?(?:(\d{1,2}/\d{1,2}/\d{2,4})\s+)?\$?\s*([\d\.,]+)",
        text,
        flags=re.I,
    ):
        amt = ars(m.group(4))
        if amt:
            found.append(
                {
                    "kind": "echeq",
                    "number": m.group(1),
                    "issuedAt": iso(m.group(2)),
                    "paidAt": iso(m.group(3)),
                    "amount": amt,
                }
            )
    for m in re.finditer(
        r"(\d{3,})\s+(\d{1,2}/\d{1,2}/\d{2,4})\s+(?:(\d{1,2}/\d{1,2}/\d{2,4})\s+)?\$?\s*([\d\.,]+)",
        text,
    ):
        amt = ars(m.group(4))
        if not amt:
            continue
        if any(x["number"] == m.group(1) for x in found):
            continue
        found.append(
            {
                "kind": "echeq",
                "number": m.group(1),
                "issuedAt": iso(m.group(2)),
                "paidAt": iso(m.group(3)),
                "amount": amt,
            }
        )
    for m in re.finditer(
        r"ENDOSO\s+ECHEQ\s+\$?\s*([\d\.,]+)",
        text,
        flags=re.I,
    ):
        amt = ars(m.group(1))
        if amt:
            found.append({"kind": "echeq", "number": None, "issuedAt": None, "paidAt": None, "amount": amt})
    for m in re.finditer(
        r"TRANSFEREN\w*[^\d$]*\$?\s*([\d\.,]+)",
        text,
        flags=re.I,
    ):
        amt = ars(m.group(1))
        if amt:
            found.append({"kind": "transfer", "number": None, "issuedAt": None, "paidAt": None, "amount": amt})
    if not found:
        upper = text.upper()
        if "ECHEQ" in upper or "ENDOSO" in upper:
            found.append({"kind": "echeq", "number": None, "issuedAt": None, "paidAt": None, "amount": None})
        elif "TRANSFER" in upper:
            found.append({"kind": "transfer", "number": None, "issuedAt": None, "paidAt": None, "amount": None})
    return found


def main():
    wb = openpyxl.load_workbook(SRC, data_only=False)
    wbv = openpyxl.load_workbook(SRC, data_only=True)
    gestion = wb[wb.sheetnames[0]]
    gv = wbv[wbv.sheetnames[0]]
    fact = wb[wb.sheetnames[1]]
    factv = wbv[wbv.sheetnames[1]]
    gestion_lines = []

    orders = {}
    clients = {}
    vendors = {}
    last = {"ord": None, "cli": None, "mes": None, "anio": None}

    for r in range(3, gestion.max_row + 1):
        ordn = fill(last["ord"], gestion.cell(r, 1).value)
        cli = fill(last["cli"], gestion.cell(r, 2).value)
        mes = fill(last["mes"], gestion.cell(r, 3).value)
        anio = fill(last["anio"], gestion.cell(r, 4).value)
        last = {"ord": ordn, "cli": cli, "mes": mes, "anio": anio}
        if not isinstance(ordn, (int, float)) or not cli:
            continue
        key = str(int(ordn))
        order = orders.setdefault(
            key,
            {
                "number": key,
                "client": str(cli).strip(),
                "month": int(mes or 1),
                "year": int(anio or 2026),
                "issuedAt": None,
                "items": [],
                "purchases": [],
                "productions": [],
                "saleInvoices": [],
            },
        )
        clients.setdefault(order["client"], {"name": order["client"], "legalName": None})

        elem = txt(gestion.cell(r, 5).value)
        item = None
        if elem:
            item = {
                "element": elem,
                "location": txt(gestion.cell(r, 6).value),
                "quantity": num(gestion.cell(r, 7).value) or 0,
                "startsAt": iso(gestion.cell(r, 8).value),
                "endsAt": iso(gestion.cell(r, 9).value),
            }
            order["items"].append(item)

        buy = parse_comp(gestion.cell(r, 13).value)
        buy_ref = None
        if buy:
            vname = vendor_name(gestion.cell(r, 14).value)
            if vname:
                vendors[vname] = {"name": vname, "kind": 0, "paymentDays": 60}
            neto = num(gestion.cell(r, 15).value) or 0
            vat = num(gv.cell(r, 16).value)
            if vat is None:
                vat = round(neto * 0.21, 2)
            diego = num(gv.cell(r, 19).value)
            if diego is None:
                diego = num(gestion.cell(r, 19).value)
            if diego is None and vname and "OSP" in vname:
                diego = round(abs(neto) * 0.1, 2) if neto else 0
            issued = iso(gestion.cell(r, 12).value)
            buy_ref = {
                **buy,
                "vendor": vname,
                "issuedAt": issued,
                "net": neto,
                "vat": round(float(vat), 2),
                "iibb": num(gv.cell(r, 17).value) or num(gestion.cell(r, 17).value) or 0,
                "percVat": num(gv.cell(r, 18).value) or num(gestion.cell(r, 18).value) or 0,
                "diegoFee": diego or 0,
                "payStatus": paid(gestion.cell(r, 21).value),
                "poNumber": txt(gestion.cell(r, 11).value) or key,
            }
            order["purchases"].append(buy_ref)
            if issued and not order["issuedAt"]:
                order["issuedAt"] = issued

        prod = parse_comp(gestion.cell(r, 27).value)
        prod_ref = None
        if prod or txt(gestion.cell(r, 28).value):
            vname = vendor_name(gestion.cell(r, 28).value)
            if vname:
                vendors[vname] = {"name": vname, "kind": 1, "paymentDays": 60}
            neto = num(gestion.cell(r, 29).value) or 0
            issued = iso(gestion.cell(r, 26).value)
            vat = num(gv.cell(r, 30).value)
            if vat is None:
                vat = num(gestion.cell(r, 30).value)
            if vat is None:
                vat = round(neto * 0.21, 2)
            if prod or neto:
                prod_ref = {
                    **(prod or {"raw": None, "docType": "A", "pos": 0, "number": 0, "credit": False}),
                    "vendor": vname,
                    "issuedAt": issued,
                    "net": neto,
                    "vat": round(float(vat), 2),
                    "payStatus": paid(gestion.cell(r, 32).value or gestion.cell(r, 33).value),
                    "poNumber": txt(gestion.cell(r, 25).value) or key,
                }
                order["productions"].append(prod_ref)

        sale = parse_comp(gestion.cell(r, 37).value)
        sale_ref = None
        if sale:
            neto = num(gestion.cell(r, 38).value) or 0
            issued = iso(gestion.cell(r, 36).value)
            status = (txt(gestion.cell(r, 45).value) or "").upper()
            vat = num(gv.cell(r, 39).value)
            if vat is None:
                vat = num(gestion.cell(r, 39).value)
            if vat is None:
                vat = round(neto * 0.21, 2)
            sale_ref = {
                **sale,
                "issuedAt": issued,
                "net": neto,
                "vat": round(float(vat), 2),
                "retVat": num(gv.cell(r, 40).value) or num(gestion.cell(r, 40).value) or 0,
                "retSuss": num(gv.cell(r, 41).value) or num(gestion.cell(r, 41).value) or 0,
                "retGan": num(gv.cell(r, 42).value) or num(gestion.cell(r, 42).value) or 0,
                "retIibb": num(gv.cell(r, 43).value) or num(gestion.cell(r, 43).value) or 0,
                "collectStatus": 1 if "COBRADO" in status else 0,
                "receiptRef": txt(gestion.cell(r, 47).value),
            }
            order["saleInvoices"].append(sale_ref)
            if issued and not order["issuedAt"]:
                order["issuedAt"] = issued

        if item or buy_ref or prod_ref or sale_ref:
            gestion_lines.append(
                {
                    "order": key,
                    "sort": r,
                    "element": item["element"] if item else None,
                    "location": item["location"] if item else None,
                    "quantity": item["quantity"] if item else 0,
                    "startsAt": item["startsAt"] if item else None,
                    "endsAt": item["endsAt"] if item else None,
                    "purchase": (
                        {"docType": buy_ref["docType"], "pos": buy_ref["pos"], "number": buy_ref["number"]}
                        if buy_ref
                        else None
                    ),
                    "production": (
                        {"docType": prod_ref["docType"], "pos": prod_ref["pos"], "number": prod_ref["number"]}
                        if prod_ref
                        else None
                    ),
                    "sale": (
                        {"docType": sale_ref["docType"], "pos": sale_ref["pos"], "number": sale_ref["number"]}
                        if sale_ref
                        else None
                    ),
                }
            )

    receipts = {}

    for r in range(2, fact.max_row + 1):
        alias = txt(fact.cell(r, 1).value)
        if not alias or alias.startswith("TOTAL"):
            continue
        legal = txt(fact.cell(r, 5).value)
        key = oc_key(fact.cell(r, 2).value)
        if not key:
            key = "SAU" if "SAU" in alias.upper() else alias.upper()
        issued = iso(fact.cell(r, 3).value) or iso(factv.cell(r, 3).value)
        client_name = resolve_client(alias, clients, orders, key if key in orders else None)
        if client_name:
            clients.setdefault(client_name, {"name": client_name, "legalName": legal})
            if legal and not clients[client_name].get("legalName"):
                clients[client_name]["legalName"] = legal
        if legal:
            for name, c in clients.items():
                if name == alias or name.startswith(alias + " ") or name.startswith(alias + "(") or alias in name:
                    if not c["legalName"]:
                        c["legalName"] = legal
        if key not in orders:
            orders[key] = {
                "number": key,
                "client": client_name or alias,
                "month": month_from_key(key, issued),
                "year": 2026,
                "issuedAt": issued,
                "items": [],
                "purchases": [],
                "productions": [],
                "saleInvoices": [],
            }
            clients.setdefault(orders[key]["client"], {"name": orders[key]["client"], "legalName": legal})

        sale = parse_comp(fact.cell(r, 4).value)
        if not sale:
            sale = {"raw": None, "docType": "A", "pos": 0, "number": 90000 + r, "credit": False}

        neto = num(fact.cell(r, 7).value) or 0
        vat = num(factv.cell(r, 8).value)
        if vat is None:
            vat = round(neto * 0.21, 2)
        collected = num(factv.cell(r, 10).value)
        if collected is None:
            collected = num(fact.cell(r, 10).value)
        receipt_ref = txt(fact.cell(r, 13).value)
        extra = {
            "legalName": legal,
            "detail": txt(fact.cell(r, 6).value),
            "collected": collected or 0,
            "receiptRef": receipt_ref,
            "retGan": num(fact.cell(r, 14).value) or 0,
            "retVat": num(fact.cell(r, 15).value) or 0,
            "retSuss": num(fact.cell(r, 16).value) or 0,
            "retIibb": num(fact.cell(r, 17).value) or 0,
            "echeq": num(fact.cell(r, 18).value) or 0,
            "bank": num(factv.cell(r, 19).value) or num(fact.cell(r, 19).value) or 0,
            "issuedAt": issued,
            "net": neto,
            "vat": round(float(vat), 2),
            "collectStatus": 1 if (collected or 0) > 0.009 else 0,
        }
        matched = False
        for inv in orders[key]["saleInvoices"]:
            if sale["number"] < 90000 and inv.get("number") == sale["number"]:
                inv.update({k: v for k, v in extra.items() if v not in (None, 0) or k.startswith("ret") or k in ("collected", "bank")})
                inv["collectStatus"] = extra["collectStatus"] or inv.get("collectStatus") or 0
                matched = True
                break
        if not matched:
            orders[key]["saleInvoices"].append({**sale, **extra})
        if issued and not orders[key]["issuedAt"]:
            orders[key]["issuedAt"] = issued

        rec_n = parse_receipt_number(receipt_ref)
        if rec_n and client_name:
            rec = receipts.setdefault(
                rec_n,
                {
                    "number": rec_n,
                    "client": client_name,
                    "issuedAt": iso(fact.cell(r, 12).value) or issued,
                    "amount": 0,
                    "invoices": [],
                },
            )
            rec["amount"] += extra["collected"] or (neto + extra["vat"])
            rec["invoices"].append({"docType": sale["docType"], "pos": sale["pos"], "number": sale["number"]})

    pay = wb[wb.sheetnames[3]]
    payment_rows = []
    last_pay = {"op": None, "medio": None, "estado": None}
    for r in range(2, pay.max_row + 1):
        prov = vendor_name(pay.cell(r, 3).value)
        raw_fc = pay.cell(r, 2).value
        importe = num(pay.cell(r, 4).value)
        if not prov and not raw_fc and importe is None:
            continue
        op = parse_op_number(fill(last_pay["op"], pay.cell(r, 8).value))
        medio = fill(last_pay["medio"], pay.cell(r, 5).value)
        estado = fill(last_pay["estado"], pay.cell(r, 7).value)
        last_pay = {
            "op": pay.cell(r, 8).value or last_pay["op"],
            "medio": pay.cell(r, 5).value or last_pay["medio"],
            "estado": pay.cell(r, 7).value or last_pay["estado"],
        }
        if not prov or importe is None:
            continue
        vendors.setdefault(prov, {"name": prov, "kind": 0, "paymentDays": 60})
        comp = parse_comp(raw_fc)
        if not comp:
            if isinstance(raw_fc, (int, float)):
                comp = {"raw": str(raw_fc), "docType": "A", "pos": 0, "number": int(raw_fc), "credit": False}
            elif txt(raw_fc) and "SALDO" in str(raw_fc).upper():
                comp = {"raw": txt(raw_fc), "docType": "NC", "pos": 0, "number": 90000 + r, "credit": True}
            else:
                comp = {"raw": txt(raw_fc), "docType": "A", "pos": 0, "number": 90000 + r, "credit": importe < 0}
        payment_rows.append(
            {
                **comp,
                "opNumber": op,
                "vendor": prov,
                "issuedAt": iso(pay.cell(r, 1).value),
                "net": importe,
                "vat": 0,
                "notes": txt(medio) if isinstance(medio, str) else None,
                "payStatus": 1 if paid(estado) else 0,
            }
        )

    lots = {}
    for row in payment_rows:
        opn = row.get("opNumber")
        if not opn:
            continue
        key = (opn, row["vendor"])
        lot = lots.setdefault(
            key,
            {
                "number": opn,
                "vendor": row["vendor"],
                "issuedAt": row["issuedAt"],
                "notes": row.get("notes"),
                "invoices": [],
            },
        )
        if row.get("notes") and not lot.get("notes"):
            lot["notes"] = row["notes"]
        if row.get("issuedAt") and (not lot.get("issuedAt") or row["issuedAt"] > lot["issuedAt"]):
            lot["issuedAt"] = row["issuedAt"]
        lot["invoices"].append(row)
    for lot in lots.values():
        lot["amount"] = round(sum(i["net"] for i in lot["invoices"]), 2)
        treasury = parse_treasury_notes(lot.get("notes"))
        if treasury and all(t.get("amount") is None for t in treasury):
            treasury = [{**t, "amount": lot["amount"]} for t in treasury]
        lot["treasury"] = treasury

    payload = {
        "clients": list(clients.values()),
        "vendors": list(vendors.values()),
        "orders": list(orders.values()),
        "gestionLines": gestion_lines,
        "receipts": list(receipts.values()),
        "paymentLots": list(lots.values()),
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(
        f"Wrote {OUT} · {len(payload['orders'])} orders · {len(payload['clients'])} clients · "
        f"{len(payload['vendors'])} vendors · {len(payload['gestionLines'])} filas · "
        f"{len(payload['receipts'])} recibos · {len(payload['paymentLots'])} OP"
    )


if __name__ == "__main__":
    main()
